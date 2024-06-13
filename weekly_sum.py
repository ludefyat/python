import pandas as pd
import os
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from yahoo_fin import stock_info as si
from parameters import *
import requests

def load_data(ticker, start_date, interval):
    #load data until 1st day of next month
    end_date = datetime(start_date.year + (start_date.month // 12), (start_date.month + 1) % 12, 1)
    si_df = pd.DataFrame()
    try:
        si_df = si.get_data(ticker, start_date=start_date - timedelta(days=1), end_date=end_date, interval=interval)
    except requests.exceptions.RequestException as e:
        print(f"Network-related error occurred: {e}")
        si_df = None
    except Exception as e:
        print(f"An error occurred: {e}\nFailed in getting {interval} data, {start_date - timedelta(days=1)} ~ {end_date}")
    else:
        si_df.dropna(inplace=True)
    return si_df


def setup_worksheet(worksheet, title):
    if title:
        worksheet.title = title
    worksheet.row_dimensions[1].height = 40
    for i, row in enumerate(worksheet.iter_rows(min_row=1, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column)):
        for cell in row:
            cell.font = Font(size=16, bold=True) if i == 0 else Font(size=15)
            cell.alignment = Alignment(vertical='center', wrapText=True)
    
    # Adjust cell to suit for content
    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                if max_line_length > max_length:
                    max_length = max_line_length
            except:
                pass
        adjusted_width = (max_length + 4) * 1.2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


def write_df_dict_to_excel(ws_dict, file_name):
    wb = Workbook()
    for key, value in ws_dict.items():
        df = value.reset_index()
        list_value = [list(df.columns)]
        list_value.extend(list(df.values))
        ws = wb.create_sheet()
        for row_index, row_data in enumerate(list_value, start=1):
            for col_index, col_data in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=col_data)
        setup_worksheet(ws, title=key)
    wb.remove(wb.active)  # Remove the default sheet created by Workbook
    wb.save(file_name)


if not DIRS_CFG['dash']:
    raise ValueError(f"dash dir NOT configured")
os.system('clear')
directory = os.getcwd()
pattern = r"\d{4}-\d{2}"
contents = [content for content in os.listdir(directory) if os.path.isdir(os.path.join(directory, content))]
matched_directories = [content for content in contents if re.match(pattern, content)]
if not matched_directories:
    print(f"Found no matched directory, quit...")
    os._exit(1)
sub_dir_list = sorted(matched_directories)
print("matched dirs found:", sub_dir_list)

stock_names = {}
for item in TICKER_CFG:
    stock_names[item["name"]] = item["ticker"] 
print(f"{stock_names}")

dash_dict = {}  #dict to store dashboard work book
chl_str = ''.join([word[0] for word in PREDICT_COLUMNS])
xls_dashboard = os.path.join(DIRS_CFG['dash'], f"dash{chl_str}_{sub_dir_list[0]}_{sub_dir_list[-1]}.xlsx")
if not os.path.exists(xls_dashboard) and len(sub_dir_list) > 1:
    xls_dashboard = os.path.join(DIRS_CFG['dash'], f"dash{chl_str}_{sub_dir_list[0]}_{sub_dir_list[-2]}.xlsx")
    print(f"Latest sum file NOT found, but found previous {xls_dashboard}...")
excel_data = pd.read_excel(xls_dashboard, sheet_name=None)
for sheet_name, df in excel_data.items():
    df.set_index(df.columns[0], inplace=True)
    dash_dict[sheet_name] = df

sub_list = sub_dir_list[-1:] if len(dash_dict) else sub_dir_list
print(f"{sub_list} is processing...")

for sub_dir in sub_list:
    print(f"{sub_dir} is processing...")
    start_date = datetime.strptime(sub_dir, "%Y-%m")
    if start_date > datetime.now():
        raise ValueError(f"start_date {sub_dir} INVALID")
    for stock, ticker in stock_names.items():
        filename = f"{stock}_{ticker}_{sub_dir}_{chl_str}.csv"
        csv_filename = os.path.join(sub_dir, filename)
        if not os.path.exists(csv_filename):
            continue
        print(f"{sub_dir}: {filename} is processing...")
        df = pd.read_csv(csv_filename)
        si_month = load_data(ticker, start_date, "1mo") #load 1 monthly data of start_date
        si_week = load_data(ticker, start_date, "1wk") #load all weekly data in the month of start_date
        if si_month is None or si_week is None:
            os._exit(100)
        print(f"month data\n{si_month.tail(7)}\n week data\n{si_week.tail(7)}")
        if si_month.empty and si_week.empty:
            print(f"{stock} data EMPTY")
            continue
        real_dict = {col + "_real": [] for col in PREDICT_COLUMNS}
        i = 0
        for column_name, column_data in df.items():
            i = i + 1
            if column_name == 'parameter':
                for j, row in enumerate(column_data):
                    date_str = row.split('\n')[0]
                    start = date_str[:10]
                    fld_str = re.findall(r'\((.*?)\)', date_str)[0]
                    interval = fld_str[1:]
                    if interval == 'month':
                        mon_start = pd.Timestamp(datetime.strptime(start, "%Y-%m-%d"))
                        if ticker[-3:] in [".SS",".SZ",".HK"]:  #adjust CST timezone to EDT
                            mon_start -= timedelta(days=1)
                        row_latest = si_month.loc[mon_start] if mon_start in si_month.index else None
                    elif interval == 'week' and start in si_week.index:
                        row_latest = si_week.loc[start]
                    else:
                        row_latest = None
                    for feature in real_dict.keys():
                        fld_val = None if row_latest is None else round(row_latest[feature[:-5]], 2) 
                        real_dict[feature].append(fld_val)
            elif column_name in PREDICT_COLUMNS:
                fld_str = f"{column_name}_real"
                if fld_str in df.keys():
                    df[fld_str] = real_dict[fld_str]
                else:
                    df.insert(i, fld_str, real_dict[fld_str])
                    i = i + 1

        df.set_index(df.columns[0], inplace=True)
        if stock in dash_dict:
            df = dash_dict[stock].combine_first(df)
        dash_dict[stock] = df
        print(f"{sub_dir}: {filename} merged into dict......\n")

xls_dashboard = os.path.join(DIRS_CFG['dash'], f"dash{chl_str}_{sub_dir_list[0]}_{sub_dir_list[-1]}.xlsx")
write_df_dict_to_excel(ws_dict=dash_dict, file_name=xls_dashboard)
print(f"{xls_dashboard} successfully produced!")


    



