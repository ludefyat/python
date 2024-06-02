import pandas as pd
import os
import csv
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from yahoo_fin import stock_info as si
from parameters import *
from load_data import end_date_si

def csv_to_dict_list(csv_file):
    data = []
    with open(csv_file, 'r', newline='') as file:
        reader = csv.reader(file)
        keys = next(reader)
        for row in reader:
            d = {}
            for i, key in enumerate(keys):
                value = row[i].strip()
                if value:
                    d[key] = value
            data.append(d)
    return data


def list_files(directory):
    for root, dirs, files in os.walk(directory):
        for filename in files:
            yield filename

def load_data(ticker, start_date, interval):
    end_date=start_date
    if interval == "1wk":
        end_date = start_date.replace(month=(start_date.month + 1)%12, day=1) 
    si_df = si.get_data(ticker, start_date=start_date - timedelta(days=1), end_date=end_date, interval=interval)
    si_df.dropna(inplace=True)
    return si_df


def setup_worksheet(worksheet, title):
    if title:
        worksheet.title = title
    worksheet.row_dimensions[1].height = 40
    for i, row in enumerate(worksheet.iter_rows(min_row=1, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column)):
        for cell in row:
            cell.font = Font(size=16, bold=True) if i == 0 else Font(size=15)
            cell.alignment = Alignment(vertical='center', wrapText=True)
    
    # Adjust cell to suit for content
    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                if max_line_length > max_length:
                    max_length = max_line_length
            except:
                pass
        adjusted_width = (max_length + 4) * 1.2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


def write_dict_to_excel(ws_dict, file_name):
    wb = Workbook()
    for key, value in ws_dict.items():
        #ws = wb.create_sheet(title=key)
        ws = wb.create_sheet()
        for row_index, row_data in enumerate(value, start=1):
            for col_index, col_data in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=col_data)
        setup_worksheet(ws, title=key)
    wb.remove(wb.active)  # Remove the default sheet created by Workbook
    wb.save(file_name)


os.system('clear')
directory = os.getcwd()
pattern = r"\d{4}-\d{2}"
contents = [content for content in os.listdir(directory) if os.path.isdir(os.path.join(directory, content))]
matched_directories = [content for content in contents if re.match(pattern, content)]
if not matched_directories:
    print(f"Found no matched directory, quit...")
    os.abort()
sub_dir_list = sorted(matched_directories)
print("matched dirs found:", sub_dir_list)

stock_names = {}
for item in TICKER_CFG:
    stock_names[item["name"]] = item["ticker"] 
print(f"{stock_names}")

dash_dict = {}  #dict to store dashboard work book
chl_str = ''.join([word[0] for word in PREDICT_COLUMNS])
for sub_dir in sub_dir_list:
    for filename in list_files(sub_dir):
        mainname, extension = os.path.splitext(filename)
        fld_lst = mainname.split('_')
        if extension != ".csv" or len(fld_lst) < 4 or fld_lst[0] not in stock_names.keys() or stock_names[fld_lst[0]] != fld_lst[1] or fld_lst[3][-len(chl_str):] != chl_str:
            print(f"{filename} ignored...")
            continue
        print(f"csv: {filename} is processing...")
        stock = fld_lst[0]
        start_date = datetime.strptime(sub_dir, "%Y-%m")
        if start_date > datetime.now():
            raise ValueError(f"start_date {sub_dir} INVALID")
        csv_filename = os.path.join(sub_dir, filename)
        df = pd.read_csv(csv_filename)
        si_month = load_data(stock_names[stock], start_date, "1mo") #load 1 monthly data of start_date
        si_week = load_data(stock_names[stock], start_date, "1wk") #load all weekly data in the month of start_date
        print(f"si_month\n{si_month.tail(7)}\n si_week\n{si_week.tail(7)}")
        if si_month.empty and si_week.empty:
            print(f"{stock} data EMPTY")
            continue
        #input()
        
        real_dict = {col + "_real": [] for col in PREDICT_COLUMNS}
        i = 0
        for column_name, column_data in df.items():
            i = i + 1
            if column_name == 'parameter':
                for j, row in enumerate(column_data):
                    date_str = row.split('\n')[0]
                    start = date_str[:10]
                    fld_str = re.findall(r'\((.*?)\)', date_str)[0]
                    interval = fld_str[1:]
                    if interval == 'month':
                        mon_start = pd.Timestamp((datetime.strptime(start, "%Y-%m-%d") - timedelta(days=1)).date())
                        row_latest = si_month.loc[mon_start] if mon_start in si_month.index else None
                    elif interval == 'week' and start in si_week.index:
                        row_latest = si_week.loc[start]
                    else:
                        row_latest = None
                    for feature in real_dict.keys():
                        fld_val = None if row_latest is None else round(row_latest[feature[:-5]], 2) 
                        real_dict[feature].append(fld_val)
            elif column_name in PREDICT_COLUMNS:
                fld_str = f"{column_name}_real"
                if fld_str in df.keys():
                    df[fld_str] = real_dict[fld_str]
                else:
                    df.insert(i, fld_str, real_dict[fld_str])
                    i = i + 1

        if stock in dash_dict:
            dash_dict[stock].extend(list(df.values))
        else:
            dash_dict[stock] = []
            dash_dict[stock].append(list(df.columns))
            dash_dict[stock].extend(list(df.values))
        print(f"{filename} merged into dict......\n")
    print(f"ALL files in {sub_dir} summarized into dict......\n")
    #input()

if DIRS_CFG['dash']:
    xls_dashboard = os.path.join(DIRS_CFG['dash'], f"dash{chl_str}_{sub_dir_list[0]}_{sub_dir_list[-1]}.xlsx")
    write_dict_to_excel(ws_dict=dash_dict, file_name=xls_dashboard)
    print(f"{xls_dashboard} successfully produced!")


    



